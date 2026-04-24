"""
excel_writer.py — Export validated records to a styled Excel workbook.

Layout:
  - One worksheet per unique School value
  - Columns: School | SEM | Degree | Course Code | Course Title | Month | Year
  - Sort order: Semester → Degree → Year → Month (chronological)
  - Header row is styled (bold, background colour)
  - Column widths auto-fitted
  - Invalid records are written to a separate "_Flagged" sheet
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.models.schema import (
    QuestionPaperRecord,
    DegreeType,
    MONTH_ORDER,
    DEGREE_ORDER,
)
from backend.utils.logger import get_logger

logger = get_logger("excel_writer")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

COLUMNS = ["School", "SEM", "Degree", "Course Code", "Course Title", "Month", "Year", "Source File"]

HEADER_FILL  = "1E3A5F"   # dark navy
HEADER_FONT  = "FFFFFF"   # white
ALT_ROW_FILL = "EAF0FB"   # light blue-grey for alternating rows

SEMESTER_ORDER = {
    "I": 1, "II": 2, "III": 3, "IV": 4,
    "V": 5, "VI": 6, "VII": 7, "VIII": 8,
    "1": 1, "2": 2, "3": 3, "4": 4,
    "5": 5, "6": 6, "7": 7, "8": 8,
}


# ---------------------------------------------------------------------------
# Sorting helpers
# ---------------------------------------------------------------------------

def _sem_key(sem: str) -> int:
    return SEMESTER_ORDER.get((sem or "").upper().strip(), 99)


def _degree_key(deg: str) -> int:
    try:
        dt = DegreeType(deg)
        return DEGREE_ORDER.get(dt, 99)
    except ValueError:
        return 99


def _month_key(month: str) -> int:
    return MONTH_ORDER.get(month, 13)


def _sort_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    df = df.copy()
    df["_sem_key"]    = df["SEM"].apply(_sem_key)
    df["_degree_key"] = df["Degree"].apply(_degree_key)
    df["_month_key"]  = df["Month"].apply(_month_key)
    df["_year_key"]   = pd.to_numeric(df["Year"], errors="coerce").fillna(9999)
    df = df.sort_values(
        by=["_sem_key", "_degree_key", "_year_key", "_month_key"],
        kind="stable",
    )
    return df.drop(columns=["_sem_key", "_degree_key", "_month_key", "_year_key"])


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def write_excel(
    valid_records: List[QuestionPaperRecord],
    invalid_records: List[QuestionPaperRecord],
    output_path: str | Path | None = None,
) -> Path:
    """
    Write all records to an Excel file.

    Parameters
    ----------
    valid_records   : list of validated QuestionPaperRecord objects
    invalid_records : list of records that failed validation
    output_path     : destination path; auto-generated if None

    Returns
    -------
    Path to the generated .xlsx file.
    """
    if output_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = Path("output") / f"OCRExtract_{ts}.xlsx"

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    logger.info("Writing Excel → %s", output_path)

    # Build a flat DataFrame
    rows = [r.to_excel_row() for r in valid_records]
    df_all = pd.DataFrame(rows, columns=COLUMNS) if rows else pd.DataFrame(columns=COLUMNS)

    # Build flagged DataFrame
    flagged_rows = [r.to_excel_row() | {"Flags": ", ".join(r.flags)} for r in invalid_records]
    df_flagged = (
        pd.DataFrame(flagged_rows, columns=COLUMNS + ["Flags"])
        if flagged_rows
        else pd.DataFrame(columns=COLUMNS + ["Flags"])
    )

    # Group by school
    schools = sorted(df_all["School"].dropna().unique()) if not df_all.empty else []
    if not schools:
        schools = ["All"]
        df_all["School"] = "All"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for school in schools:
            sheet_name = _safe_sheet_name(school)
            df_school = df_all[df_all["School"] == school].copy()
            df_school = _sort_df(df_school)
            df_school.to_excel(writer, sheet_name=sheet_name, index=False)

        if not df_flagged.empty:
            df_flagged.to_excel(writer, sheet_name="_Flagged", index=False)

    # Apply styling with openpyxl
    _style_workbook(output_path, schools)

    logger.info(
        "Excel written: %d valid rows, %d flagged rows, %d sheets.",
        len(valid_records), len(invalid_records), len(schools),
    )
    return output_path


# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------

def _style_workbook(path: Path, schools: List[str]) -> None:
    wb = load_workbook(path)

    sheet_names = [_safe_sheet_name(s) for s in schools]
    if "_Flagged" in wb.sheetnames:
        sheet_names.append("_Flagged")

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        _style_sheet(ws)

    wb.save(path)


def _style_sheet(ws) -> None:
    """Apply header styling, alternating row fills, and auto column widths."""
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    alt_fill    = PatternFill("solid", fgColor=ALT_ROW_FILL)
    header_font = Font(bold=True, color=HEADER_FONT, name="Calibri", size=11)
    body_font   = Font(name="Calibri", size=10)
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    col_widths: dict[int, int] = {}

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            col = cell.column
            # Track max content width
            val_len = len(str(cell.value)) if cell.value is not None else 0
            col_widths[col] = max(col_widths.get(col, 0), val_len)

            if row_idx == 1:
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = center_align
            else:
                cell.font      = body_font
                cell.border    = thin_border
                cell.alignment = (
                    center_align
                    if col in (1, 2, 3, 4, 6, 7)   # School/SEM/Degree/Code/Month/Year
                    else left_align
                )
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

    # Set column widths (min 10, max 50)
    for col_idx, max_len in col_widths.items():
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = max(10, min(max_len + 4, 50))

    # Freeze header row
    ws.freeze_panes = "A2"


def _safe_sheet_name(name: str) -> str:
    """Truncate to 31 chars and strip illegal Excel sheet-name characters."""
    import re
    name = re.sub(r"[\\/*?:\[\]]", "_", name)
    return name[:31] if name else "Sheet"
